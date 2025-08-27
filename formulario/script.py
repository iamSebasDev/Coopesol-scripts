import pandas as pd
from openpyxl import load_workbook

# --- Configuración ---
ruta_excel = "plantilla.xlsx"
ruta_csv = "datos.csv"

# 1. Cargar datos del CSV y AGRUPAR por NO y MES
df = pd.read_csv("/home/trabajo/Proyectos/Coopesol/Coopesol-scripts/formulario/datos.csv")
print("Datos originales del CSV:")
print(df)

# Agrupar y sumar las transacciones duplicadas
df_agrupado = df.groupby(['NO', 'MES']).agg({
    'CAPITAL': 'sum',
    'INTERES': 'sum', 
    'MORA': 'sum'
}).reset_index()

print("\nDatos agrupados y sumados:")
print(df_agrupado)

# 2. Abrir el Excel existente
wb = load_workbook("/home/trabajo/Proyectos/Coopesol/Coopesol-scripts/formulario/plantilla.xlsx")
ws = wb["Hoja 1"]

# Mapear cada mes a sus columnas
columnas = {
    "sep-24": {"capital": 22, "interes": 23, "mora": 24},
    "oct-24": {"capital": 25, "interes": 26, "mora": 27},
    "nov-24": {"capital": 28, "interes": 29, "mora": 30},
    "dic-24": {"capital": 31, "interes": 32, "mora": 33},
    "ene-25": {"capital": 34, "interes": 35, "mora": 36},
    "feb-25": {"capital": 37, "interes": 38, "mora": 39},
    "mar-25": {"capital": 40, "interes": 41, "mora": 42},
}

# 3. Rellenar datos con los valores SUMADOS
for index, row in df_agrupado.iterrows():
    numero_asociado = row["NO"]
    mes = row["MES"].lower().strip()
    capital = row["CAPITAL"]
    interes = row["INTERES"] 
    mora = row["MORA"]

    print(f"Buscando asociado NO: {numero_asociado} - Mes: {mes}")
    print(f"Valores a insertar - Capital: {capital}, Interés: {interes}, Mora: {mora}")

    # Buscar la fila por número de asociado (columna D)
    encontrado = False
    for fila in range(2, ws.max_row + 1):
        no_excel = ws.cell(row=fila, column=4).value  # Columna D (NO.)
        if no_excel and str(no_excel).strip() == str(numero_asociado):
            print(f"✓ Encontrado NO. {numero_asociado} en fila {fila}")
            # Insertar valores SUMADOS
            ws.cell(row=fila, column=columnas[mes]["capital"], value=capital)
            ws.cell(row=fila, column=columnas[mes]["interes"], value=interes)
            ws.cell(row=fila, column=columnas[mes]["mora"], value=mora)
            encontrado = True
            break
    
    if not encontrado:
        print(f"✗ NO encontrado: {numero_asociado}")

# 4. Guardar
wb.save("plantillaEnero2.xlsx")
print("\n✅ Excel actualizado con los datos SUMADOS del CSV")